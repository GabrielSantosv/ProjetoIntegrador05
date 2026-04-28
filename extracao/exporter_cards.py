"""
exporter_cards.py
Adicione este arquivo em extracao/exporter_cards.py

Gera o Excel com 3 abas:
  - resumo     : card visual por arquivo (igual às imagens)
  - pages      : dados brutos por página
  - tables     : tabelas extraídas

Uso:
    from extracao.exporter_cards import exportar_com_cards
    exportar_com_cards(result, Path("03_OUTPUT/analise_consolidada.xlsx"))
"""
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

import pandas as pd

# ─────────────────────────────────────────────────────────────
# Paleta de cores
# ─────────────────────────────────────────────────────────────
COR = {
    # fundos de cabeçalho por nível de risco
    "maximo_bg":      "FEE2E2",   # vermelho claro
    "maximo_fg":      "991B1B",
    "medio_bg":       "FEF3C7",   # amarelo claro
    "medio_fg":       "92400E",
    "informativo_bg": "D1FAE5",   # verde claro
    "informativo_fg": "065F46",
    "desconhecido_bg":"F3F4F6",
    "desconhecido_fg":"374151",
    # status
    "nada_constar":   "059669",   # verde
    "constar":        "D97706",   # âmbar
    # cabeçalhos de seção
    "secao_bg":       "F9FAFB",
    "secao_borda":    "E5E7EB",
    # processo
    "proc_num":       "1D4ED8",
    # alerta
    "alerta_bg":      "FFFBEB",
    "alerta_borda":   "FDE68A",
    "alerta_fg":      "78350F",
    # cabeçalho da tabela pages/tables
    "header_bg":      "1E3A5F",
    "header_fg":      "FFFFFF",
    # linha alternada
    "linha_par":      "F8FAFC",
}

TIPO_DESCRICAO = {
    "civel_estadual":    "Certidão Estadual de Distribuições Cíveis — TJSP",
    "criminal_estadual": "Certidão Estadual de Distribuições Criminais — TJSP",
    "cnd_federal":       "Certidão Negativa de Débitos Federais — RFB/PGFN",
    "cnd_estadual":      "Certidão Negativa de Débitos Estaduais — SEFAZ-SP",
    "trf3":              "Certidão Judicial — Tribunal Regional Federal",
    "civel_federal":     "Certidão Judicial Cível — CJF (multi-região)",
    "criminal_federal":  "Certidão Judicial Criminal — CJF (multi-região)",
    "cndt":              "Certidão Negativa de Débitos Trabalhistas — TST",
    "ceat":              "Certidão de Ações Trabalhistas",
    "eleitoral":         "Certidão para Fins Eleitorais",
    "desconhecido":      "Tipo não identificado pelo classificador",
}

# ─────────────────────────────────────────────────────────────
# Helpers de estilo
# ─────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10, italic=False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Arial")

def _border_box(color="E5E7EB") -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _border_bottom(color="E5E7EB") -> Border:
    s = Side(style="thin", color=color)
    return Border(bottom=s)

def _alinhar(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _escrever(ws, row, col, valor, bold=False, color="000000",
              size=10, bg=None, align="left", wrap=False,
              italic=False, border=None) -> None:
    cell = ws.cell(row=row, column=col, value=valor)
    cell.font = _font(bold=bold, color=color, size=size, italic=italic)
    cell.alignment = _alinhar(h=align, v="center", wrap=wrap)
    if bg:
        cell.fill = _fill(bg)
    if border:
        cell.border = border

def _mesclar(ws, row1, col1, row2, col2) -> None:
    ws.merge_cells(
        start_row=row1, start_column=col1,
        end_row=row2,   end_column=col2,
    )

# ─────────────────────────────────────────────────────────────
# Extração de dados auxiliares do raw_text
# ─────────────────────────────────────────────────────────────

def _extrair_rg(raw: str) -> str:
    m = re.search(r"RG[:\s]+([0-9.\-X]+)", raw, re.IGNORECASE)
    return m.group(1) if m else ""

def _extrair_nascimento(raw: str) -> str:
    m = re.search(r"nascido\s+em\s+(\d{2}/\d{2}/\d{4})", raw, re.IGNORECASE)
    local = re.search(r"natural\s+de\s+([A-ZÀ-Ü][a-zA-Zà-ü\s]+ - [A-Z]{2})", raw)
    if m and local:
        return f"{m.group(1)} — {local.group(1)}"
    if m:
        return m.group(1)
    return ""

def _extrair_certidao_num(raw: str) -> str:
    m = re.search(r"CERTIDÃO\s+N[Oº°][\.:]\s*([\d]+)", raw, re.IGNORECASE)
    if m:
        return m.group(1)
    m2 = re.search(r"Nº\s+([\d]+/\d+)", raw)
    return m2.group(1) if m2 else ""

def _processos_geo(meta: dict) -> list[str]:
    raw = meta.get("processes_geometric", [])
    if isinstance(raw, list):
        return raw
    try:
        return json.loads(str(raw).replace("'", '"'))
    except Exception:
        return []

def _contexto_processo(proc: str, raw: str) -> str:
    idx = raw.find(proc)
    if idx == -1:
        return ""
    trecho = raw[max(0, idx - 80): idx + 150]
    partes = []
    foro = re.search(r"Foro\s+(?:de\s+)?([A-ZÀ-Ü][a-zA-Zà-ü\s]+)\s*[-–]\s*(\d+ª?\s*Vara[^.·\n]*)", trecho)
    acao = re.search(r"(?:Ação|Crime|Inquérito|Falência)[^.·\n]{5,60}", trecho, re.IGNORECASE)
    data = re.search(r"Data:\s*(\d{2}/\d{2}/\d{4})", trecho)
    reqte = re.search(r"Reqte:\s*([^\n·*]{3,40})", trecho)
    if foro:
        partes.append(f"Foro {foro.group(1).strip()} – {foro.group(2).strip()}")
    if acao:
        partes.append(acao.group(0).strip())
    if data:
        partes.append(f"Data: {data.group(1)}")
    if reqte:
        partes.append(f"Reqte: {reqte.group(1).strip()}")
    return " · ".join(partes)

def _alertas(tipo: str, status: str | None, tem_homonimos: bool,
             total_paginas: int, raw: str) -> list[str]:
    alertas = []
    if tipo == "desconhecido":
        alertas.append("Tipo não identificado. Verifique a pasta ou o conteúdo do PDF.")
    if "trf" in tipo or "civel_federal" in tipo or "criminal_federal" in tipo:
        alertas.append(
            "Caso especial: padrão 'NÃO CONSTAM' (CJF) — diferente do padrão TJSP. "
            "Verifique o NADA_CONSTAR_REGEX em parser.py."
        )
    if tem_homonimos:
        alertas.append(
            "Documento com dois blocos: CPF qualificado (NADA CONSTAR) + "
            "homônimos (CONSTAR). Pipeline deve encerrar após o primeiro bloco."
        )
    if tipo in ("civel_estadual",) and "FALÊNCIA" in raw.upper():
        alertas.append(
            "Certidão de falência: classifier.py não detecta subtipo explicitamente — "
            "classificado como civel_estadual genérico."
        )
    if status == "NADA CONSTAR" and not tem_homonimos and total_paginas == 1:
        alertas.append(
            "Certidão totalmente negativa. Parser encontra NADA CONSTAR e interrompe "
            "na página 1. Nenhum processo a extrair."
        )
    return alertas

# ─────────────────────────────────────────────────────────────
# Aba RESUMO — card por arquivo
# ─────────────────────────────────────────────────────────────

def _aba_resumo(ws, records: list) -> None:
    # Larguras das colunas A–H
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 3

    # Agrupa por arquivo
    grupos: dict[str, list] = {}
    for r in records:
        fname = Path(r.source_file).name
        grupos.setdefault(fname, []).append(r)

    linha = 2  # começa na linha 2 (margem)

    for filename, recs in grupos.items():
        rep = recs[0]
        tipo       = rep.document_type or "desconhecido"
        risco      = rep.nivel_risco   or "informativo"
        status     = rep.status
        raw        = rep.raw_text or ""
        total_pag  = len(recs)
        meta       = rep.metadata or {}

        todos_geo: list[str] = []
        for r in recs:
            todos_geo.extend(_processos_geo(r.metadata or {}))

        tem_homonimos   = bool(todos_geo) and status == "NADA CONSTAR"
        is_neg_pura     = status == "NADA CONSTAR" and not tem_homonimos
        cert_num        = _extrair_certidao_num(raw)
        descricao       = TIPO_DESCRICAO.get(tipo, tipo)
        folhas          = f"{total_pag} {'folha' if total_pag == 1 else 'folhas'}"
        cor_risco       = COR.get(f"{risco}_bg", COR["desconhecido_bg"])
        cor_risco_fg    = COR.get(f"{risco}_fg", COR["desconhecido_fg"])

        # ── Linha 1: nome do arquivo + badge ──────────────────────
        _mesclar(ws, linha, 2, linha, 5)
        _escrever(ws, linha, 2, filename,
                  bold=True, size=12, bg="FFFFFF",
                  border=_border_box("E5E7EB"))
        ws.row_dimensions[linha].height = 22

        label_badge = f"{tipo.replace('_', ' ')} · risco {risco}" if tipo != "desconhecido" else "desconhecido"
        if is_neg_pura:
            label_badge = f"{tipo.replace('_', ' ')} · NEGATIVA PURA"
        _mesclar(ws, linha, 6, linha, 7)
        _escrever(ws, linha, 6, label_badge,
                  bold=True, size=9,
                  color=cor_risco_fg, bg=cor_risco,
                  align="center",
                  border=_border_box(cor_risco_fg))
        linha += 1

        # ── Linha 2: subtítulo ────────────────────────────────────
        _mesclar(ws, linha, 2, linha, 7)
        subtitulo = descricao
        if cert_num:
            subtitulo += f"   ·   Certidão nº {cert_num}"
        subtitulo += f"   ·   {folhas}"
        _escrever(ws, linha, 2, subtitulo,
                  size=9, color="6B7280", bg="FFFFFF",
                  border=_border_bottom())
        ws.row_dimensions[linha].height = 16
        linha += 1

        # ── Caixas: TITULAR | STATUS ──────────────────────────────
        ws.row_dimensions[linha].height = 14

        _escrever(ws, linha, 2, "TITULAR" + (" (QUALIFICADO)" if status == "NADA CONSTAR" else ""),
                  bold=True, size=8, color="9CA3AF",
                  bg=COR["secao_bg"], border=_border_box(COR["secao_borda"]))
        _escrever(ws, linha, 4, "STATUS CERTIDÃO" if tem_homonimos else "STATUS",
                  bold=True, size=8, color="9CA3AF",
                  bg=COR["secao_bg"], border=_border_box(COR["secao_borda"]))
        _mesclar(ws, linha, 4, linha, 5)
        linha += 1

        # Titular — dados
        nome = rep.name or "—"
        cpf  = rep.cpf  or "—"
        rg   = _extrair_rg(raw)
        nasc = _extrair_nascimento(raw)
        linhas_titular = [nome, f"CPF: {cpf}"]
        if rg:
            linhas_titular.append(f"RG: {rg}")
        if nasc:
            linhas_titular.append(f"Nasc: {nasc}")

        # Status — dados
        if status == "NADA CONSTAR" and tem_homonimos:
            linhas_status = [
                ("NADA CONSTAR (CPF qualificado)", COR["nada_constar"]),
                ("CONSTAR (homônimos não qualificados)", COR["constar"]),
            ]
        elif status == "NADA CONSTAR":
            linhas_status = [
                ("NADA CONSTAR", COR["nada_constar"]),
                ("Sem bloco de homônimos", "6B7280"),
                (f"Certidão de {total_pag} folha{'s' if total_pag > 1 else ''}", "6B7280"),
            ]
        else:
            linhas_status = [
                (status or "—", COR["constar"]),
            ]

        max_linhas = max(len(linhas_titular), len(linhas_status))
        for i in range(max_linhas):
            ws.row_dimensions[linha].height = 15
            if i < len(linhas_titular):
                _escrever(ws, linha, 2, linhas_titular[i],
                          size=10, bold=(i == 0),
                          bg=COR["secao_bg"],
                          border=_border_box(COR["secao_borda"]))
                _mesclar(ws, linha, 2, linha, 3)
            if i < len(linhas_status):
                txt, cor_st = linhas_status[i]
                _escrever(ws, linha, 4, txt,
                          size=10, bold=(i == 0), color=cor_st,
                          bg=COR["secao_bg"],
                          border=_border_box(COR["secao_borda"]))
                _mesclar(ws, linha, 4, linha, 7)
            linha += 1

        linha += 1  # espaço

        # ── Processos geométricos ─────────────────────────────────
        if todos_geo:
            label_geo = "PROCESSOS DETECTADOS VIA MARCADOR »"
            if tem_homonimos:
                label_geo += "  (HOMÔNIMOS — NÃO QUALIFICADOS)"

            ws.row_dimensions[linha].height = 14
            _mesclar(ws, linha, 2, linha, 7)
            _escrever(ws, linha, 2, label_geo,
                      bold=True, size=8, color="9CA3AF",
                      bg=COR["secao_bg"],
                      border=_border_box(COR["secao_borda"]))
            linha += 1

            for proc in todos_geo[:15]:
                ctx = _contexto_processo(proc, raw)
                ws.row_dimensions[linha].height = 15
                _escrever(ws, linha, 2, f"» {proc}",
                          size=10, color=COR["proc_num"],
                          bg=COR["secao_bg"],
                          border=_border_box(COR["secao_borda"]))
                _mesclar(ws, linha, 2, linha, 3)
                _escrever(ws, linha, 4, ctx,
                          size=9, color="6B7280", italic=True,
                          bg=COR["secao_bg"],
                          border=_border_box(COR["secao_borda"]))
                _mesclar(ws, linha, 4, linha, 7)
                linha += 1

            if len(todos_geo) > 15:
                ws.row_dimensions[linha].height = 14
                _mesclar(ws, linha, 2, linha, 7)
                _escrever(ws, linha, 2,
                          f"+ {len(todos_geo) - 15} processos adicionais (ver aba 'pages')",
                          size=9, color="6B7280", italic=True,
                          bg=COR["secao_bg"])
                linha += 1

            linha += 1

        # ── Estrutura TRF multi-região ────────────────────────────
        if ("trf" in tipo or "federal" in tipo) and total_pag > 1:
            ws.row_dimensions[linha].height = 14
            _mesclar(ws, linha, 2, linha, 7)
            _escrever(ws, linha, 2,
                      "ESTRUTURA: 1 CERTIDÃO = N REGIÕES (PÁGINAS SEPARADAS POR TRIBUNAL)",
                      bold=True, size=8, color="9CA3AF",
                      bg=COR["secao_bg"],
                      border=_border_box(COR["secao_borda"]))
            linha += 1

            regioes_map = {
                "1": "TRF 1ª Região — bases: PJe, SEEU, JEF Virtual, Proc. Digital",
                "2": "TRF 2ª Região — Eproc ES, RJ, TRF2",
                "3": "TRF 3ª Região — SAIP 1º/2º grau, PJe",
                "5": "TRF 5ª Região — PJE-AL/CE/PB/PE/RN/SE/T5, TEBAS, ESPARTA",
            }
            for reg, desc in regioes_map.items():
                if f"{reg}ª REGIÃO" in raw.upper():
                    ws.row_dimensions[linha].height = 14
                    _mesclar(ws, linha, 2, linha, 7)
                    _escrever(ws, linha, 2, f"· {desc}",
                              size=9, color="374151",
                              bg=COR["secao_bg"],
                              border=_border_box(COR["secao_borda"]))
                    linha += 1

            cod = re.search(r"Código de validação:\s*([A-Z0-9.]+)", raw)
            if cod:
                ws.row_dimensions[linha].height = 14
                _mesclar(ws, linha, 2, linha, 7)
                _escrever(ws, linha, 2, f"Código de validação: {cod.group(1)}",
                          size=9, color="374151",
                          bg=COR["secao_bg"],
                          border=_border_box(COR["secao_borda"]))
                linha += 1

            linha += 1

        # ── Alertas ───────────────────────────────────────────────
        alertas = _alertas(tipo, status, tem_homonimos, total_pag, raw)
        for alerta in alertas:
            ws.row_dimensions[linha].height = 30
            _mesclar(ws, linha, 2, linha, 7)
            _escrever(ws, linha, 2, alerta,
                      size=9, color=COR["alerta_fg"],
                      bg=COR["alerta_bg"],
                      border=_border_box(COR["alerta_borda"]),
                      wrap=True)
            linha += 1

        linha += 2  # espaço entre cards

# ─────────────────────────────────────────────────────────────
# Aba PAGES — dados brutos formatados
# ─────────────────────────────────────────────────────────────

def _aba_pages(ws, records: list) -> None:
    colunas = [
        ("Arquivo",              "source_file",        28),
        ("Página",               "page_number",         7),
        ("Tipo",                 "document_type",       18),
        ("Risco",                "nivel_risco",         12),
        ("Status",               "status",              14),
        ("Nome",                 "name",                22),
        ("CPF",                  "cpf",                 16),
        ("Nº Processo",          "process_number",      26),
        ("Data",                 "date",                12),
        ("Valor",                "value",               12),
        ("Tipo de Ação",         "tipo_acao",           20),
        ("Sit. Processual",      "situacao_processual", 20),
        ("Vara",                 "vara",                18),
        ("Foro",                 "foro",                18),
        ("Fonte Texto",          "text_source",         12),
    ]

    # Cabeçalho
    for col_idx, (label, _, width) in enumerate(colunas, 1):
        letra = get_column_letter(col_idx)
        ws.column_dimensions[letra].width = width
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = _font(bold=True, color=COR["header_fg"], size=10)
        cell.fill      = _fill(COR["header_bg"])
        cell.alignment = _alinhar(h="center")
        cell.border    = _border_box("FFFFFF")

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # Dados
    COR_RISCO = {
        "maximo":      ("FEE2E2", "991B1B"),
        "medio":       ("FEF3C7", "92400E"),
        "informativo": ("D1FAE5", "065F46"),
    }

    for row_idx, r in enumerate(records, 2):
        bg = COR["linha_par"] if row_idx % 2 == 0 else "FFFFFF"
        risco = getattr(r, "nivel_risco", "informativo")
        meta = r.metadata or {}

        for col_idx, (_, field, _) in enumerate(colunas, 1):
            if field == "text_source":
                valor = meta.get("text_source", "")
            else:
                valor = getattr(r, field, None)
                if field == "source_file":
                    valor = Path(str(valor)).name

            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.font      = _font(size=9)
            cell.alignment = _alinhar(v="center")
            cell.border    = _border_box("E5E7EB")

            # Colore coluna de risco
            if field == "nivel_risco" and risco in COR_RISCO:
                bg_r, fg_r = COR_RISCO[risco]
                cell.fill = _fill(bg_r)
                cell.font = _font(size=9, color=fg_r, bold=True)
            elif field == "status":
                if valor == "NADA CONSTAR":
                    cell.font = _font(size=9, color=COR["nada_constar"], bold=True)
                elif valor == "POSITIVA":
                    cell.font = _font(size=9, color=COR["constar"], bold=True)
                cell.fill = _fill(bg)
            else:
                cell.fill = _fill(bg)

        ws.row_dimensions[row_idx].height = 15

    # Filtro automático
    ws.auto_filter.ref = f"A1:{get_column_letter(len(colunas))}1"

# ─────────────────────────────────────────────────────────────
# Aba TABLES — tabelas extraídas
# ─────────────────────────────────────────────────────────────

def _aba_tables(ws, tables: list) -> None:
    colunas = [
        ("Arquivo",       "source_file",  28),
        ("Página",        "page_number",   7),
        ("Índice Tabela", "table_index",   12),
        ("Tipo Doc.",     "document_type", 18),
        ("Células (JSON)","cells_json",    60),
    ]

    for col_idx, (label, _, width) in enumerate(colunas, 1):
        letra = get_column_letter(col_idx)
        ws.column_dimensions[letra].width = width
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = _font(bold=True, color=COR["header_fg"], size=10)
        cell.fill      = _fill(COR["header_bg"])
        cell.alignment = _alinhar(h="center")

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    for row_idx, t in enumerate(tables, 2):
        bg = COR["linha_par"] if row_idx % 2 == 0 else "FFFFFF"
        vals = {
            "source_file":   Path(str(t.source_file)).name,
            "page_number":   t.page_number,
            "table_index":   t.table_index,
            "document_type": t.document_type,
            "cells_json":    json.dumps(t.cells, ensure_ascii=False) if t.cells else "",
        }
        for col_idx, (_, field, _) in enumerate(colunas, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=vals.get(field))
            cell.font      = _font(size=9)
            cell.fill      = _fill(bg)
            cell.alignment = _alinhar(v="center",
                                      wrap=(field == "cells_json"))
            cell.border    = _border_box("E5E7EB")
        ws.row_dimensions[row_idx].height = 15 if not t.cells else 30

# ─────────────────────────────────────────────────────────────
# Função pública principal
# ─────────────────────────────────────────────────────────────

def exportar_com_cards(result: Any, output_path: Path) -> Path:
    """
    Gera o Excel com 3 abas:
      resumo  — card visual por arquivo (igual às imagens)
      pages   — dados brutos por página com formatação
      tables  — tabelas extraídas

    Parâmetros:
        result      : PipelineResult (com .records e .tables)
        output_path : Path do arquivo .xlsx de saída

    Retorno:
        Path do arquivo gerado
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # Aba 1: resumo (renomeia a aba padrão)
    ws_resumo = wb.active
    ws_resumo.title = "resumo"
    ws_resumo.sheet_view.showGridLines = False
    _aba_resumo(ws_resumo, result.records)

    # Aba 2: pages
    ws_pages = wb.create_sheet("pages")
    _aba_pages(ws_pages, result.records)

    # Aba 3: tables
    ws_tables = wb.create_sheet("tables")
    _aba_tables(ws_tables, result.tables)

    wb.save(output_path)
    return output_path